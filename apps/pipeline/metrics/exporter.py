"""
MetricsExporter — exports soa_metrics_results to a multi-sheet xlsx file.

Reads from the DB (never recomputes). Produces one sheet per slice dimension
plus an overall summary and an optional trend sheet.
"""
import logging
import os
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

import soa_shared.config as config
from soa_shared.database import engine

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

_HDR_FILL      = PatternFill("solid", fgColor="1F4E79")   # dark blue header
_HDR_FONT      = Font(color="FFFFFF", bold=True)
_PRIMARY_FILL  = PatternFill("solid", fgColor="D6E4F7")   # light blue — primary merchant
_GREEN_FILL    = PatternFill("solid", fgColor="C6EFCE")
_YELLOW_FILL   = PatternFill("solid", fgColor="FFEB9C")
_RED_FILL      = PatternFill("solid", fgColor="FFC7CE")

# Number formats
_FMT_PCT   = "0.0%"
_FMT_IDX   = "0.000"
_FMT_INT   = "#,##0"
_FMT_PCT2  = "0.00%"


def _soa_fill(soa_pct: Optional[float]) -> Optional[PatternFill]:
    if soa_pct is None:
        return None
    if soa_pct >= 0.30:
        return _GREEN_FILL
    if soa_pct >= 0.15:
        return _YELLOW_FILL
    return _RED_FILL


def _write_header(ws, columns: List[str]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autofit(ws) -> None:
    """Estimate column widths based on max content length."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 8), 40)


def _fmt(value: Any, fmt: str) -> Any:
    """Return value as-is (formatting applied via cell.number_format)."""
    return value


class MetricsExporter:

    def __init__(self, cycle_code: str) -> None:
        self.cycle_code = cycle_code
        self._cycle_id: Optional[int] = None

    def export(self, output_path: Optional[str] = None) -> str:
        """
        Build and write the xlsx workbook.
        Returns the full path of the written file.
        """
        if output_path is None:
            exports_dir = getattr(config, "SOA_EXPORTS_DIR", "/soa/exports")
            os.makedirs(exports_dir, exist_ok=True)
            output_path = os.path.join(
                exports_dir, f"soa_metrics_{self.cycle_code}.xlsx"
            )

        self._cycle_id = self._load_cycle_id()
        if self._cycle_id is None:
            raise ValueError(f"Cycle '{self.cycle_code}' not found in soa_cycles.")

        wb = Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        self._sheet_overall(wb)
        self._sheet_by_dim(wb, "BY CATEGORY",    "category")
        self._sheet_by_dim(wb, "BY STAGE",        "stage")
        self._sheet_by_dim(wb, "BY SPECIFICITY",  "specificity")
        self._sheet_by_dim(wb, "BY PERSONA",       "persona")
        self._sheet_by_platform(wb)
        self._sheet_trend(wb)

        wb.save(output_path)
        logger.info("MetricsExporter: wrote %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # Sheet 1: OVERALL SCORES
    # ------------------------------------------------------------------

    def _sheet_overall(self, wb: Workbook) -> None:
        ws = wb.create_sheet("OVERALL SCORES")
        cols = [
            "Entity", "Entity Type", "Role",
            "Mention Rate", "SoA%", "Position Index",
            "RSI Score", "Deal Citation Rate", "Platform Dist Index",
            "Total Runs", "Total Mentions",
        ]
        _write_header(ws, cols)

        rows = self._fetch_metrics("overall", "all")
        rows.sort(key=lambda r: (r["soa_pct"] or 0), reverse=True)

        soa_col = cols.index("SoA%") + 1

        for row in rows:
            role = "Primary" if row["entity_role"] == "primary" else "Competitor"
            data = [
                row["merchant_name"],
                row.get("entity_type", "retailer"),
                role,
                row["mention_rate"],
                row["soa_pct"],
                row["position_index"],
                row["rsi_score"],
                row["deal_citation_rate"],
                row["platform_dist_index"],
                row["total_runs"],
                row["total_mentions"],
            ]
            ws.append(data)
            r = ws.max_row

            if role == "Primary":
                for cell in ws[r]:
                    cell.fill = _PRIMARY_FILL

            soa_cell = ws.cell(r, soa_col)
            fill = _soa_fill(row["soa_pct"])
            if fill:
                soa_cell.fill = fill

            self._apply_row_formats(ws, r, cols)

        _autofit(ws)

    # ------------------------------------------------------------------
    # Sheets 2-5: BY DIMENSION
    # ------------------------------------------------------------------

    def _sheet_by_dim(self, wb: Workbook, title: str, dim: str) -> None:
        ws = wb.create_sheet(title)
        dim_label = dim.replace("_", " ").title()
        cols = [
            dim_label, "Entity",
            "Mention Rate", "SoA%", "Position Index",
            "RSI Score", "Deal Citation Rate",
            "Total Runs",
        ]
        _write_header(ws, cols)

        # Fetch all slice_values for this dimension
        rows = self._fetch_all_for_dim(dim)
        # Sort by dimension value then SoA% desc
        rows.sort(key=lambda r: (r["slice_value"], -(r["soa_pct"] or 0)))

        soa_col = cols.index("SoA%") + 1

        for row in rows:
            data = [
                row["slice_value"],
                row["merchant_name"],
                row["mention_rate"],
                row["soa_pct"],
                row["position_index"],
                row["rsi_score"],
                row["deal_citation_rate"],
                row["total_runs"],
            ]
            ws.append(data)
            r = ws.max_row
            soa_cell = ws.cell(r, soa_col)
            fill = _soa_fill(row["soa_pct"])
            if fill:
                soa_cell.fill = fill
            self._apply_row_formats(ws, r, cols)

        _autofit(ws)

    # ------------------------------------------------------------------
    # Sheet 6: BY PLATFORM
    # ------------------------------------------------------------------

    def _sheet_by_platform(self, wb: Workbook) -> None:
        ws = wb.create_sheet("BY PLATFORM")
        cols = [
            "Platform", "Entity",
            "Mention Rate", "SoA%", "Position Index",
            "RSI Score", "Deal Citation Rate", "Total Runs",
            "Pct Runs With Search", "Mention Rate (Search)", "Mention Rate (No Search)",
        ]
        _write_header(ws, cols)

        rows = self._fetch_all_for_dim("platform")
        rows.sort(key=lambda r: (r["slice_value"], -(r["soa_pct"] or 0)))

        search_stats = self._fetch_search_triggered_stats()
        soa_col = cols.index("SoA%") + 1

        for row in rows:
            platform = row["slice_value"]
            mid = row["entity_id"]

            if platform in ("chatgpt", "claude") and mid in search_stats:
                st = search_stats[mid]
                pct_search       = st.get("pct_search")
                mr_search        = st.get("mention_rate_search")
                mr_no_search     = st.get("mention_rate_no_search")
            else:
                pct_search = mr_search = mr_no_search = None

            data = [
                platform,
                row["merchant_name"],
                row["mention_rate"],
                row["soa_pct"],
                row["position_index"],
                row["rsi_score"],
                row["deal_citation_rate"],
                row["total_runs"],
                pct_search,
                mr_search,
                mr_no_search,
            ]
            ws.append(data)
            r = ws.max_row
            soa_cell = ws.cell(r, soa_col)
            fill = _soa_fill(row["soa_pct"])
            if fill:
                soa_cell.fill = fill
            self._apply_row_formats(ws, r, cols)

        _autofit(ws)

    # ------------------------------------------------------------------
    # Sheet 7: TREND
    # ------------------------------------------------------------------

    def _sheet_trend(self, wb: Workbook) -> None:
        ws = wb.create_sheet("TREND")
        cols = [
            "Cycle", "Merchant", "SoA%",
            "Mention Rate", "RSI Score", "Deal Citation Rate",
        ]
        _write_header(ws, cols)

        sql = """
            SELECT
                c.cycle_code,
                e.name AS merchant_name,
                mr.soa_pct,
                mr.mention_rate,
                mr.rsi_score,
                mr.deal_citation_rate
            FROM soa_metrics_results mr
            JOIN soa_cycles   c ON c.id = mr.cycle_id
            JOIN soa_entities e ON e.id = mr.entity_id
            WHERE mr.slice_type  = 'overall'
              AND mr.slice_value = 'all'
              AND c.status       = 'complete'
            ORDER BY c.cycle_code ASC, mr.soa_pct DESC NULLS LAST
        """
        with engine.connect() as conn:
            trend_rows = conn.execute(text(sql)).fetchall()

        cycles = {r[0] for r in trend_rows}
        if len(cycles) < 2:
            ws.append(["Trend data available after 2+ complete cycles."])
            return

        soa_col = cols.index("SoA%") + 1
        for row in trend_rows:
            ws.append([row[0], row[1], row[2], row[3], row[4], row[5]])
            r = ws.max_row
            fill = _soa_fill(row[2])
            if fill:
                ws.cell(r, soa_col).fill = fill
            self._apply_row_formats(ws, r, cols)

        _autofit(ws)

    # ------------------------------------------------------------------
    # Data fetching helpers
    # ------------------------------------------------------------------

    def _load_cycle_id(self) -> Optional[int]:
        with engine.connect() as conn:
            row = conn.execute(
                text("SELECT id FROM soa_cycles WHERE cycle_code = :cc"),
                {"cc": self.cycle_code},
            ).fetchone()
        return int(row[0]) if row else None

    def _fetch_metrics(
        self, slice_type: str, slice_value: str
    ) -> List[Dict]:
        sql = """
            SELECT
                mr.entity_id,
                e.name                              AS merchant_name,
                e.slug                              AS merchant_slug,
                COALESCE(e.entity_type, 'retailer') AS entity_type,
                COALESCE(ce.role, 'competitor')     AS entity_role,
                mr.total_runs,
                mr.total_mentions,
                mr.mention_rate,
                mr.soa_pct,
                mr.position_index,
                mr.rsi_score,
                mr.deal_citation_rate,
                mr.platform_dist_index
            FROM soa_metrics_results mr
            JOIN soa_entities e ON e.id = mr.entity_id
            LEFT JOIN soa_cycle_entities ce
                ON ce.entity_id = mr.entity_id AND ce.cycle_id = mr.cycle_id
            WHERE mr.cycle_id    = :cycle_id
              AND mr.slice_type  = :slice_type
              AND mr.slice_value = :slice_value
        """
        with engine.connect() as conn:
            rows = conn.execute(
                text(sql),
                {
                    "cycle_id":    self._cycle_id,
                    "slice_type":  slice_type,
                    "slice_value": slice_value,
                },
            ).fetchall()
        return [self._row_to_dict(r) for r in rows]

    def _fetch_all_for_dim(self, dim: str) -> List[Dict]:
        sql = """
            SELECT
                mr.entity_id,
                e.name                              AS merchant_name,
                e.slug                              AS merchant_slug,
                COALESCE(e.entity_type, 'retailer') AS entity_type,
                COALESCE(ce.role, 'competitor')     AS entity_role,
                mr.slice_value,
                mr.total_runs,
                mr.total_mentions,
                mr.mention_rate,
                mr.soa_pct,
                mr.position_index,
                mr.rsi_score,
                mr.deal_citation_rate,
                mr.platform_dist_index
            FROM soa_metrics_results mr
            JOIN soa_entities e ON e.id = mr.entity_id
            LEFT JOIN soa_cycle_entities ce
                ON ce.entity_id = mr.entity_id AND ce.cycle_id = mr.cycle_id
            WHERE mr.cycle_id   = :cycle_id
              AND mr.slice_type = :slice_type
            ORDER BY mr.slice_value, mr.soa_pct DESC NULLS LAST
        """
        with engine.connect() as conn:
            rows = conn.execute(
                text(sql),
                {"cycle_id": self._cycle_id, "slice_type": dim},
            ).fetchall()
        return [self._row_to_dict(r, has_slice_val=True) for r in rows]

    def _fetch_search_triggered_stats(self) -> Dict[int, Dict]:
        """
        For chatgpt and claude platforms, return per-merchant search-triggered stats:
          pct_search, mention_rate_search, mention_rate_no_search.
        Returns {} if no runs with search_triggered data exist.
        Aggregates across both platforms (chatgpt + claude) since both expose
        the search_triggered signal.
        """
        sql = """
            SELECT
                cm.entity_id,
                COUNT(*)                                                                        AS total_runs,
                SUM(CASE WHEN r.search_triggered = TRUE  THEN 1 ELSE 0 END)                   AS search_runs,
                SUM(CASE WHEN r.search_triggered = FALSE THEN 1 ELSE 0 END)                   AS no_search_runs,
                SUM(CASE WHEN r.search_triggered = TRUE  AND cm.mentioned THEN 1 ELSE 0 END)  AS search_mentions,
                SUM(CASE WHEN r.search_triggered = FALSE AND cm.mentioned THEN 1 ELSE 0 END)  AS no_search_mentions
            FROM soa_runs r
            JOIN soa_coded_mentions cm ON cm.run_id = r.id
            WHERE r.cycle_id = :cycle_id
              AND r.platform  IN ('chatgpt', 'claude')
              AND r.status    = 'success'
              AND cm.entity_id IS NOT NULL
            GROUP BY cm.entity_id
        """
        with engine.connect() as conn:
            rows = conn.execute(text(sql), {"cycle_id": self._cycle_id}).fetchall()

        result: Dict[int, Dict] = {}
        for row in rows:
            mid, total, search, no_search, s_mentions, ns_mentions = (
                int(row[0]), int(row[1] or 0), int(row[2] or 0),
                int(row[3] or 0), int(row[4] or 0), int(row[5] or 0),
            )
            result[mid] = {
                "pct_search":             round(search / total, 4) if total > 0 else None,
                "mention_rate_search":    round(s_mentions / search, 4) if search > 0 else None,
                "mention_rate_no_search": round(ns_mentions / no_search, 4) if no_search > 0 else None,
            }
        return result

    def _row_to_dict(self, row, has_slice_val: bool = False) -> Dict:
        if has_slice_val:
            return {
                "entity_id":           int(row[0]),
                "merchant_name":       row[1],
                "merchant_slug":       row[2],
                "entity_type":         row[3],
                "entity_role":         row[4],
                "slice_value":         str(row[5]) if row[5] is not None else "",
                "total_runs":          int(row[6] or 0),
                "total_mentions":      int(row[7] or 0),
                "mention_rate":        row[8],
                "soa_pct":             row[9],
                "position_index":      row[10],
                "rsi_score":           row[11],
                "deal_citation_rate":  row[12],
                "platform_dist_index": row[13],
            }
        else:
            return {
                "entity_id":           int(row[0]),
                "merchant_name":       row[1],
                "merchant_slug":       row[2],
                "entity_type":         row[3],
                "entity_role":         row[4],
                "total_runs":          int(row[5] or 0),
                "total_mentions":      int(row[6] or 0),
                "mention_rate":        row[7],
                "soa_pct":             row[8],
                "position_index":      row[9],
                "rsi_score":           row[10],
                "deal_citation_rate":  row[11],
                "platform_dist_index": row[12],
            }

    # ------------------------------------------------------------------
    # Number formatting
    # ------------------------------------------------------------------

    def _apply_row_formats(self, ws, row_idx: int, cols: List[str]) -> None:
        """Apply number formats to each cell based on the column name."""
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row_idx, col_idx)
            if any(k in col_name for k in ("Rate", "SoA%", "Pct", "Search)")):
                cell.number_format = _FMT_PCT
            elif any(k in col_name for k in ("Index", "RSI")):
                cell.number_format = _FMT_IDX
            elif col_name in ("Total Runs", "Total Mentions"):
                cell.number_format = _FMT_INT
